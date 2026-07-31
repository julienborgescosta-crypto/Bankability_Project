from pathlib import Path

import openpyxl
import pytest

from core.models import ProjectInputs

SAMPLE_DATA_DIR = Path(__file__).resolve().parent.parent / "sample_data"


@pytest.fixture
def simple_inputs() -> ProjectInputs:
    """3 annees, chiffres ronds choisis pour etre verifiables a la main :
    construction (CAPEX -1000) puis 2 annees d'exploitation identiques
    (Revenue 500, OPEX -100, TURPE -20 -> CFADS 380)."""
    return ProjectInputs(
        name="Test Project",
        location="Testville",
        segment="HTB1",
        cod="2026-01-01",
        operating_years=2,
        usable_power_mw=10.0,
        usable_energy_mwh=20.0,
        capex_initial_keur=1000.0,
        capex_repowering_keur=0.0,
        repowering=False,
        opex_year1_keur=100.0,
        opex_adjustment_keur=0.0,
        turpe_fixed_eur_per_kw=0.0,
        years=[2025, 2026, 2027],
        capex_keur=[-1000.0, 0.0, 0.0],
        opex_keur=[0.0, -100.0, -100.0],
        end_of_life_keur=[0.0, 0.0, 0.0],
        revenues_keur=[0.0, 500.0, 500.0],
        turpe_keur=[0.0, -20.0, -20.0],
        net_cashflow_keur=[-1000.0, 380.0, 380.0],
        wacc=0.10,
        gearing_pct=0.5,
        interest_rate=0.05,
        debt_tenor_years=2,
    )


@pytest.fixture
def ramp_up_inputs() -> ProjectInputs:
    """4 annees : construction, une annee de ramp-up SANS CAPEX ni revenu (COD
    tombe apres le dernier decaissement CAPEX), puis 2 annees d'exploitation.
    Sert a verifier que le service de la dette ne demarre pas pendant le ramp-up."""
    return ProjectInputs(
        name="Ramp-up Project",
        location="Testville",
        segment="HTB1",
        cod="2027-01-01",
        operating_years=2,
        usable_power_mw=10.0,
        usable_energy_mwh=20.0,
        capex_initial_keur=1000.0,
        capex_repowering_keur=0.0,
        repowering=False,
        opex_year1_keur=100.0,
        opex_adjustment_keur=0.0,
        turpe_fixed_eur_per_kw=0.0,
        years=[2025, 2026, 2027, 2028],
        capex_keur=[-1000.0, 0.0, 0.0, 0.0],
        opex_keur=[0.0, -10.0, -100.0, -100.0],
        end_of_life_keur=[0.0, 0.0, 0.0, 0.0],
        revenues_keur=[0.0, 0.0, 500.0, 500.0],
        turpe_keur=[0.0, 0.0, -20.0, -20.0],
        net_cashflow_keur=[-1000.0, -10.0, 380.0, 380.0],
        wacc=0.10,
        gearing_pct=0.5,
        interest_rate=0.05,
        debt_tenor_years=2,
    )


@pytest.fixture
def repowering_inputs() -> ProjectInputs:
    """11 annees avec 2 sorties de CAPEX (construction 2025 puis repowering 2031),
    chacune avec ses propres termes de financement - sert a verifier la logique
    de dette a 2 tranches independantes de financial_engine.compute_results."""
    years = list(range(2025, 2036))  # 2025..2035
    capex = [-1000.0] + [0.0] * 5 + [-600.0] + [0.0] * 4
    # CFADS annuel = 380 (Revenue 500, OPEX -100, TURPE -20) sur toutes les annees
    # d'exploitation ; l'annee de repowering (2031) est en pause (0 partout).
    revenues = [0.0] + [500.0] * 5 + [0.0] + [500.0] * 4
    opex = [0.0] + [-100.0] * 5 + [0.0] + [-100.0] * 4
    turpe = [0.0] + [-20.0] * 5 + [0.0] + [-20.0] * 4
    return ProjectInputs(
        name="Repowering Project",
        location="Testville",
        segment="HTB1",
        cod="2026-01-01",
        operating_years=10,
        usable_power_mw=10.0,
        usable_energy_mwh=20.0,
        capex_initial_keur=1000.0,
        capex_repowering_keur=600.0,
        repowering=True,
        opex_year1_keur=100.0,
        opex_adjustment_keur=0.0,
        turpe_fixed_eur_per_kw=0.0,
        years=years,
        capex_keur=capex,
        opex_keur=opex,
        end_of_life_keur=[0.0] * len(years),
        revenues_keur=revenues,
        turpe_keur=turpe,
        net_cashflow_keur=[
            c + o + t + r for c, o, t, r in zip(capex, opex, turpe, revenues, strict=True)
        ],
        wacc=0.10,
        gearing_pct=0.5,
        interest_rate=0.05,
        debt_tenor_years=3,
        repowering_gearing_pct=0.6,
        repowering_interest_rate=0.08,
        repowering_debt_tenor_years=2,
    )


@pytest.fixture
def sample_summary_bp_path() -> Path:
    path = SAMPLE_DATA_DIR / "260612_BP_Stockage_Standalone__Claude.xlsx"
    assert path.exists(), "Fixture manquante - lancer sample_data/build_sample_xlsx.py"
    return path


@pytest.fixture
def sample_full_bp_path(tmp_path: Path) -> Path:
    """Construit un classeur synthetique minimal au format 'complet'
    (onglets O-Financials / O-Control), avec des donnees fabriquees - jamais
    les vraies donnees BP confidentielles de l'utilisateur, pour que ce test
    reste executable sans le fichier reel (qui n'est jamais commite)."""
    path = tmp_path / "synthetic_full_bp.xlsx"
    wb = openpyxl.Workbook()

    control = wb.active
    control.title = "O-Control"
    control["B3"], control["C3"] = "Name", "Synthetic Project"
    control["B4"], control["C4"] = "Location city", "Testville"
    control["B5"], control["C5"] = "Segment", "HTB1"
    control["B6"], control["C6"] = "BESS commercial operation date (COD)", "2027-01-01"
    control["B7"], control["C7"], control["D7"] = "BESS operating time", 2, "years"
    control["B8"], control["C8"], control["D8"] = "ESS Usable Power @PoC (AC)", 10, "MW"
    control["B9"], control["C9"], control["D9"] = "ESS Usable Energy @PoC BoL (AC)", 20, "MWh"
    control["G4"], control["H4"], control["I4"] = "IRR", 0.08, 0.06
    control["G12"], control["H12"] = "NPV", 150.0
    control["G17"], control["H17"] = "CAPEX (w/o DSRA and financing fees)", 1000.0
    control["G28"], control["H28"], control["I28"] = "Debt", 500.0, 0.5
    control["G42"], control["H42"] = "Maturity", 2
    control["G43"], control["H43"] = "All-in rate (fixed part)", 0.05
    control["G44"], control["H44"] = "Average DSCR", 1.4
    control["G45"], control["H45"] = "Min. DSCR", 1.4
    control["G7"], control["H7"] = "Equity discount factor", 0.12

    financials = wb.create_sheet("O-Financials")
    financials["K8"] = "Year"
    years = [2025, 2026, 2027]
    for offset, year in enumerate(years):
        financials.cell(row=8, column=12 + offset, value=year)

    def series(row: int, label: str, values: list[float]) -> None:
        financials.cell(row=row, column=4, value=label)
        for offset, value in enumerate(values):
            financials.cell(row=row, column=12 + offset, value=value)

    series(11, "Revenues", [0.0, 500.0, 500.0])
    series(16, "Operating Costs", [0.0, -100.0, -100.0])
    series(27, "Operating Taxes", [0.0, -20.0, -20.0])
    series(58, "CAPEX (w/o DSRA and financing fees)", [-1000.0, 0.0, 0.0])

    wb.save(path)
    return path


@pytest.fixture
def sample_full_bp_with_i_project_path(tmp_path: Path) -> Path:
    """Meme classeur synthetique que sample_full_bp_path, avec en plus un onglet
    I-Project reproduisant le piege du fichier reel : certains libelles ("Maturity",
    "All-in rate (fixed part)") sont repetes une fois pour la dette senior et une
    fois pour la dette de repowering - sert a tester le parametre `occurrence`."""
    path = tmp_path / "synthetic_full_bp_with_i_project.xlsx"
    wb = openpyxl.Workbook()

    control = wb.active
    control.title = "O-Control"
    control["B3"], control["C3"] = "Name", "Synthetic Project"
    control["B4"], control["C4"] = "Location city", "Testville"
    control["B5"], control["C5"] = "Segment", "HTB1"
    control["B6"], control["C6"] = "BESS commercial operation date (COD)", "2027-01-01"
    control["B7"], control["C7"], control["D7"] = "BESS operating time", 2, "years"
    control["B8"], control["C8"], control["D8"] = "ESS Usable Power @PoC (AC)", 10, "MW"
    control["B9"], control["C9"], control["D9"] = "ESS Usable Energy @PoC BoL (AC)", 20, "MWh"
    control["G4"], control["H4"], control["I4"] = "IRR", 0.08, 0.06
    control["G12"], control["H12"] = "NPV", 150.0
    control["G17"], control["H17"] = "CAPEX (w/o DSRA and financing fees)", 1000.0
    control["G28"], control["H28"], control["I28"] = "Debt", 500.0, 0.5
    control["G42"], control["H42"] = "Maturity", 2
    control["G43"], control["H43"] = "All-in rate (fixed part)", 0.05
    control["G44"], control["H44"] = "Average DSCR", 1.4
    control["G45"], control["H45"] = "Min. DSCR", 1.4
    control["G7"], control["H7"] = "Equity discount factor", 0.12

    financials = wb.create_sheet("O-Financials")
    financials["K8"] = "Year"
    years = [2025, 2026, 2027]
    for offset, year in enumerate(years):
        financials.cell(row=8, column=12 + offset, value=year)

    def series(row: int, label: str, values: list[float]) -> None:
        financials.cell(row=row, column=4, value=label)
        for offset, value in enumerate(values):
            financials.cell(row=row, column=12 + offset, value=value)

    series(11, "Revenues", [0.0, 500.0, 500.0])
    series(16, "Operating Costs", [0.0, -100.0, -100.0])
    series(27, "Operating Taxes", [0.0, -20.0, -20.0])
    series(58, "CAPEX (w/o DSRA and financing fees)", [-1000.0, 0.0, 0.0])

    # I-Project a un ordre de colonnes different de O-Control : libelle | unite |
    # valeur (ex. "Maturity | years | 10"), pas libelle | valeur | unite - piege
    # rencontre sur le fichier reel, reproduit ici intentionnellement.
    i_project = wb.create_sheet("I-Project")
    i_project["B10"], i_project["C10"], i_project["D10"] = (
        "CAPEX (w/o DSRA and financing fees)",
        "k€",
        -1050.0,
    )
    i_project["B20"], i_project["C20"], i_project["D20"] = "Target DSCR - Period 1", "x", 1.5
    i_project["B30"], i_project["C30"], i_project["D30"] = (
        "Maturity",
        "years",
        10,
    )  # senior debt (1ere occurrence)
    i_project["B31"], i_project["C31"], i_project["D31"] = (
        "All-in rate (fixed part)",
        "%",
        0.07,
    )  # senior debt
    i_project["B40"], i_project["C40"], i_project["D40"] = (
        "Gearing",
        "%",
        0.6,
    )  # repowering debt (libelle unique)
    i_project["B41"], i_project["C41"], i_project["D41"] = (
        "Maturity",
        "years",
        8,
    )  # repowering debt (2e occurrence)
    i_project["B42"], i_project["C42"], i_project["D42"] = (
        "All-in rate (fixed part)",
        "%",
        0.09,
    )  # repowering debt

    wb.save(path)
    return path
