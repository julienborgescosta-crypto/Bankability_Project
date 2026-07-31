from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import yaml

from .models import ProjectInputs

DEFAULT_MAPPING_PATH = Path(__file__).resolve().parent.parent / "config" / "bp_mapping.yaml"
DEFAULT_FULL_MAPPING_PATH = (
    Path(__file__).resolve().parent.parent / "config" / "bp_mapping_full.yaml"
)
FULL_FORMAT_SHEETS = {"O-Financials", "O-Control"}


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return " ".join(text.split())


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and value.strip() in ("", "-")


def load_grid(file_or_path, sheet_name: str | None = None) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(file_or_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    return [[cell.value for cell in row] for row in worksheet.iter_rows()]


def _find_label(grid: list[list[Any]], label: str) -> tuple[list[Any] | None, int | None]:
    target = _normalize(label)
    for row in grid:
        for idx, cell in enumerate(row):
            if _normalize(cell) == target:
                return row, idx
    for row in grid:
        for idx, cell in enumerate(row):
            norm = _normalize(cell)
            if norm and norm.startswith(target):
                return row, idx
    return None, None


def _field_spec(entry: Any) -> tuple[str, int]:
    """A mapping entry is either a plain label string (offset=1: the first non-blank
    cell to the right of the label) or {label, offset} to pick the Nth such cell -
    e.g. a row like "Debt | 19251 (k€) | 70% (gearing)" needs offset=2 for the %."""
    if isinstance(entry, dict):
        return entry["label"], int(entry.get("offset", 1))
    return entry, 1


def _scalar_value(grid: list[list[Any]], label: str, offset: int = 1) -> Any:
    row, idx = _find_label(grid, label)
    if row is None:
        return None
    seen = 0
    for cell in row[idx + 1 :]:
        if not _is_blank(cell):
            seen += 1
            if seen == offset:
                return cell
    return None


def _series_values(
    grid: list[list[Any]], label: str, start_col: int, length: int, required: bool = True
) -> list[float]:
    row, idx = _find_label(grid, label)
    if row is None:
        if required:
            raise ValueError(f"Ligne introuvable dans le BP pour le libelle '{label}'.")
        return [0.0] * length
    values = row[start_col : start_col + length]
    values = list(values) + [None] * (length - len(values))
    return [0.0 if _is_blank(v) else float(v) for v in values]


def load_mapping(mapping_path: Path = DEFAULT_MAPPING_PATH) -> dict:
    with open(mapping_path, encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def parse_summary_bp(file_or_path, mapping_path: Path = DEFAULT_MAPPING_PATH) -> ProjectInputs:
    mapping = load_mapping(mapping_path)
    grid = load_grid(file_or_path, sheet_name=mapping.get("sheet"))

    scalar_fields = mapping["scalar_fields"]
    series_fields = mapping["series_fields"]

    year_row, year_col = _find_label(grid, series_fields["years"])
    if year_row is None:
        raise ValueError(
            "Impossible de localiser la ligne 'Year' dans le fichier BP. "
            "Verifiez que le fichier correspond au format attendu (voir config/bp_mapping.yaml)."
        )

    year_values: list[int] = []
    for cell in year_row[year_col + 1 :]:
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            year_values.append(int(cell))
        else:
            break
    if not year_values:
        raise ValueError("La ligne 'Year' du BP ne contient aucune annee exploitable.")

    length = len(year_values)
    start_col = year_col + 1

    def series_of(key: str) -> list[float]:
        return _series_values(grid, series_fields[key], start_col, length)

    def scalar_of(key: str, default: Any = None) -> Any:
        value = _scalar_value(grid, scalar_fields[key])
        return default if value is None else value

    def scalar_float(key: str) -> float | None:
        value = _scalar_value(grid, scalar_fields[key])
        return None if value is None else float(value)

    repowering_raw = scalar_of("repowering", "Non")
    repowering = _normalize(repowering_raw) in ("oui", "yes", "true", "1")

    return ProjectInputs(
        name=str(scalar_of("name", "")).strip(),
        location=str(scalar_of("location", "") or "").strip(),
        segment=str(scalar_of("segment", "")).strip(),
        cod=str(scalar_of("cod", "")),
        operating_years=int(float(scalar_of("operating_years", 0) or 0)),
        usable_power_mw=float(scalar_of("usable_power_mw", 0) or 0),
        usable_energy_mwh=float(scalar_of("usable_energy_mwh", 0) or 0),
        capex_initial_keur=float(scalar_of("capex_initial_keur", 0) or 0),
        capex_repowering_keur=float(scalar_of("capex_repowering_keur", 0) or 0),
        repowering=repowering,
        opex_year1_keur=float(scalar_of("opex_year1_keur", 0) or 0),
        opex_adjustment_keur=float(scalar_of("opex_adjustment_keur", 0) or 0),
        turpe_fixed_eur_per_kw=float(scalar_of("turpe_fixed_eur_per_kw", 0) or 0),
        years=year_values,
        capex_keur=series_of("capex_keur"),
        opex_keur=series_of("opex_keur"),
        end_of_life_keur=series_of("end_of_life_keur"),
        revenues_keur=series_of("revenues_keur"),
        turpe_keur=series_of("turpe_keur"),
        net_cashflow_keur=series_of("net_cashflow_keur"),
        reported_irr=scalar_float("reported_irr"),
        reported_wacc=scalar_float("reported_wacc"),
        reported_npv_keur=scalar_float("reported_npv_keur"),
        wacc=scalar_float("reported_wacc") or 0.10,
    )


def detect_format(file_or_path) -> str:
    """ "full" = the real multi-tab model (O-Financials + O-Control present),
    "summary" = a single-tab export like the illustrative CSV/xlsx fixture."""
    workbook = openpyxl.load_workbook(file_or_path, read_only=True)
    sheets = set(workbook.sheetnames)
    workbook.close()
    if hasattr(file_or_path, "seek"):
        file_or_path.seek(0)
    return "full" if FULL_FORMAT_SHEETS.issubset(sheets) else "summary"


def parse_bp(file_or_path) -> ProjectInputs:
    fmt = detect_format(file_or_path)
    if fmt == "full":
        return parse_full_bp(file_or_path)
    return parse_summary_bp(file_or_path)


def parse_full_bp(file_or_path, mapping_path: Path = DEFAULT_FULL_MAPPING_PATH) -> ProjectInputs:
    mapping = load_mapping(mapping_path)
    scalar_grid = load_grid(file_or_path, sheet_name=mapping["scalar_sheet"])
    if hasattr(file_or_path, "seek"):
        file_or_path.seek(0)
    series_grid = load_grid(file_or_path, sheet_name=mapping["series_sheet"])

    scalar_fields = mapping["scalar_fields"]
    series_fields = mapping["series_fields"]

    years_label, _ = _field_spec(series_fields["years"])
    year_row, year_col = _find_label(series_grid, years_label)
    if year_row is None:
        raise ValueError(
            f"Impossible de localiser la ligne 'Year' dans l'onglet '{mapping['series_sheet']}'. "
            "Verifiez config/bp_mapping_full.yaml."
        )
    year_values: list[int] = []
    for cell in year_row[year_col + 1 :]:
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            year_values.append(int(cell))
        else:
            break
    if not year_values:
        raise ValueError("La ligne 'Year' ne contient aucune annee exploitable.")

    length = len(year_values)
    start_col = year_col + 1

    def series_of(key: str, required: bool = True) -> list[float]:
        label, _ = _field_spec(series_fields[key])
        return _series_values(series_grid, label, start_col, length, required=required)

    def scalar_of(key: str, default: Any = None) -> Any:
        label, offset = _field_spec(scalar_fields[key])
        value = _scalar_value(scalar_grid, label, offset)
        return default if value is None else value

    def scalar_float(key: str) -> float | None:
        label, offset = _field_spec(scalar_fields[key])
        value = _scalar_value(scalar_grid, label, offset)
        return None if value is None else float(value)

    capex_series = series_of("capex_keur")
    opex_series = series_of("opex_keur")
    turpe_series = series_of("turpe_keur")
    revenue_series = series_of("revenues_keur")
    eol_series = series_of("end_of_life_keur", required=False)
    net_cf_series = [
        c + o + t + r + e
        for c, o, t, r, e in zip(
            capex_series, opex_series, turpe_series, revenue_series, eol_series, strict=True
        )
    ]

    interest_rate = scalar_float("interest_rate")
    gearing_pct = scalar_float("gearing_pct")
    equity_discount = scalar_float("equity_discount_factor")
    tenor = scalar_of("debt_tenor_years")

    wacc = None
    if interest_rate is not None and gearing_pct is not None and equity_discount is not None:
        # No explicit WACC cell in this model - approximate as a gearing-weighted
        # blend of the real cost of debt and the equity discount factor found in
        # the sheet, rather than falling back to an arbitrary constant.
        wacc = gearing_pct * interest_rate + (1 - gearing_pct) * equity_discount

    cod_raw = scalar_of("cod", "")
    cod = cod_raw.date().isoformat() if hasattr(cod_raw, "date") else str(cod_raw)

    return ProjectInputs(
        name=str(scalar_of("name", "")).strip(),
        location=str(scalar_of("location", "") or "").strip(),
        segment=str(scalar_of("segment", "")).strip(),
        cod=cod,
        operating_years=int(float(scalar_of("operating_years", 0) or 0)),
        usable_power_mw=float(scalar_of("usable_power_mw", 0) or 0),
        usable_energy_mwh=float(scalar_of("usable_energy_mwh", 0) or 0),
        capex_initial_keur=float(scalar_of("capex_initial_keur", 0) or 0),
        capex_repowering_keur=0.0,
        repowering=any(v != 0 for v in eol_series),
        opex_year1_keur=next((abs(v) for v in opex_series if v != 0), 0.0),
        opex_adjustment_keur=0.0,
        turpe_fixed_eur_per_kw=0.0,
        years=year_values,
        capex_keur=capex_series,
        opex_keur=opex_series,
        end_of_life_keur=eol_series,
        revenues_keur=revenue_series,
        turpe_keur=turpe_series,
        net_cashflow_keur=net_cf_series,
        reported_irr=scalar_float("reported_irr"),
        reported_equity_irr=scalar_float("reported_equity_irr"),
        reported_npv_keur=scalar_float("reported_npv_keur"),
        reported_dscr_avg=scalar_float("reported_dscr_avg"),
        reported_dscr_min=scalar_float("reported_dscr_min"),
        wacc=wacc if wacc is not None else 0.10,
        gearing_pct=gearing_pct if gearing_pct is not None else 0.70,
        interest_rate=interest_rate if interest_rate is not None else 0.05,
        debt_tenor_years=int(tenor) if tenor else 15,
    )
