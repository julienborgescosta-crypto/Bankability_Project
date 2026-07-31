"""Builds a .xlsx fixture that mirrors the layout of the summary BP tab shown
in the reference CSV (260612_BP_Stockage_Standalone__Claude.csv), but with
properly typed Excel cells (numbers as numbers, correctly signed) instead of
the French-formatted display strings ("14 707", "(469)") found in that export.
Used as the Temps-1 test fixture for core/bp_parser.py.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

YEARS = list(range(2027, 2060))  # 2027..2059 inclusive, 33 years

OPEX_OP_YEARS = [469, 493, 501, 511, 515, 526, 534, 541, 547, 562, 571, 572, 582, 590, 601, 612, 638, 649, 661, 673]
REVENUE_OP_YEARS = [3087, 2572, 2373, 2305, 2216, 2148, 2061, 1986, 1987, 1932, 1894, 1863, 1845, 1788, 1812, 1782, 1764, 1722, 1715, 1674]
TURPE_OP_YEARS = [272, 265, 255, 253, 248, 245, 241, 238, 235, 232, 228, 225, 222, 219, 215, 213, 210, 272, 262, 256]

N_OP_YEARS = len(OPEX_OP_YEARS)  # 20, matches "BESS operating time" = 20 years


def _padded(values_negative: list[int]) -> list[float]:
    """1 construction year (0) + N operating years (as given) + trailing zeros to 33 years."""
    trailing = len(YEARS) - 1 - len(values_negative)
    return [0.0] + [float(v) for v in values_negative] + [0.0] * trailing


def build(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BP_Stockage_Standalone"

    def scalar(row: int, label: str, value, unit=None):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=value)
        if unit is not None:
            ws.cell(row=row, column=4, value=unit)

    def series(row: int, label: str, values: list[float]):
        ws.cell(row=row, column=2, value=label)
        for offset, value in enumerate(values):
            ws.cell(row=row, column=8 + offset, value=value)

    scalar(2, "Name", "Fougeres")
    scalar(3, "Location city", None)
    scalar(4, "Segment", "HTB2")
    scalar(5, "BESS commercial operation date (COD)", "01/01/2028")
    scalar(6, "BESS operating time", N_OP_YEARS, "years")
    scalar(7, "BESS Usable Power", 20, "MW")
    scalar(8, "BESS Usable Energy", 40, "MWh")

    ws.cell(row=10, column=7, value="Year")
    for offset, year in enumerate(YEARS):
        ws.cell(row=10, column=8 + offset, value=year)

    capex_row = [0.0] * len(YEARS)
    capex_row[0] = -14707.0
    series(12, "CAPEX", capex_row)
    series(13, "OPEX (live = C-SPV)", _padded([-v for v in OPEX_OP_YEARS]))
    series(14, "End of Life Value", [0.0] * len(YEARS))
    series(15, "Revenues", _padded([float(v) for v in REVENUE_OP_YEARS]))
    series(16, "TURPE (variable + fixe)", _padded([-v for v in TURPE_OP_YEARS]))

    net_cf_op_years = [
        rev - opex - turpe for rev, opex, turpe in zip(REVENUE_OP_YEARS, OPEX_OP_YEARS, TURPE_OP_YEARS)
    ]
    net_cf_row = _padded([float(v) for v in net_cf_op_years])
    net_cf_row[0] = -14707.0
    series(17, "Net Cashflow", net_cf_row)

    scalar(20, "IRR", 0.0666)
    scalar(22, "WACC", 0.10)
    scalar(24, "NPV", -2420)

    scalar(28, "OPEX an 1 (live C-SPV, info)", 469)
    scalar(29, "OPEX ajustement (k€, + = cout en plus)", 0)
    scalar(30, "Repowering (Oui/Non)", "Non")
    scalar(31, "CAPEX initial (live I-Project)", 14707)
    scalar(32, "CAPEX repowering (live I-Project)", 2965)
    scalar(33, "TURPE fixe - Grid charge Aurora (€/kW)", 3.83)

    wb.save(path)
    print(f"Fichier ecrit : {path}")


if __name__ == "__main__":
    build(Path(__file__).resolve().parent / "260612_BP_Stockage_Standalone__Claude.xlsx")
